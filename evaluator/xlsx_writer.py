"""Write an interactive xlsx workbook.

Every numeric input lives on the `Inputs` sheet with a defined name. All
downstream sheets (Scorecard, Org, Valuation_*) reference those named ranges
via Excel formulas, so editing an input recomputes the workbook.
"""
from typing import Dict, List
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.utils import get_column_letter

from . import schema, rubric as rb, org as org_mod, screening as sc_mod, definitions as defs, equity as eq


HEADER = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SECTION = Font(bold=True, size=12, color="1F4E78")
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _header_row(ws, row, headers, widths=None):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = HEADER
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")
        if widths and i - 1 < len(widths):
            ws.column_dimensions[get_column_letter(i)].width = widths[i - 1]


def _safe_name(key: str) -> str:
    # Excel defined names: letters, digits, underscore; can't start with digit.
    return "in_" + key


def _percent_format(field: schema.Field) -> str | None:
    if field.kind == schema.PCT:
        return "0.0%"
    if field.kind in (schema.NUM,):
        return "#,##0"
    return None


def write_workbook(inputs: Dict, results: Dict, path: str) -> None:
    wb = Workbook()

    # ---------------- Inputs sheet ----------------
    ws = wb.active
    ws.title = "Inputs"
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 60

    ws["A1"] = "Idea Evaluator — Inputs"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E78")
    ws["A2"] = "Edit yellow cells. Other sheets recompute automatically."
    ws["A2"].font = Font(italic=True, color="666666")

    row = 4
    sections_seen = set()
    addr_for_key: Dict[str, str] = {}

    for f in schema.ALL_FIELDS:
        if f.section not in sections_seen:
            ws.cell(row=row, column=1, value=f.section.upper()).font = SECTION
            row += 1
            sections_seen.add(f.section)
        ws.cell(row=row, column=1, value=f.prompt).alignment = Alignment(wrap_text=True)
        val_cell = ws.cell(row=row, column=2, value=inputs.get(f.key, f.default))
        val_cell.fill = PatternFill("solid", fgColor="FFF2CC")
        val_cell.border = BORDER
        fmt = _percent_format(f)
        if fmt:
            val_cell.number_format = fmt
        if f.help:
            ws.cell(row=row, column=3, value=f.help).font = Font(italic=True, color="808080")
        addr = f"Inputs!$B${row}"
        addr_for_key[f.key] = addr
        # Defined name for formula readability
        if f.kind in (schema.NUM, schema.INT, schema.PCT, schema.SCORE):
            dn = DefinedName(name=_safe_name(f.key), attr_text=f"Inputs!$B${row}")
            wb.defined_names[_safe_name(f.key)] = dn
        row += 1

    def nm(k: str) -> str:
        return _safe_name(k)

    # ---------------- Scorecard sheet ----------------
    sc = wb.create_sheet("Scorecard")
    sc.column_dimensions["A"].width = 28
    sc.column_dimensions["B"].width = 12
    sc.column_dimensions["C"].width = 12
    sc.column_dimensions["D"].width = 40
    sc["A1"] = "Evaluation Scorecard"
    sc["A1"].font = Font(bold=True, size=14, color="1F4E78")

    _header_row(sc, 3, ["Category", "Weight", "Avg Score (1-5)", "Sub-scores"])
    r = 4
    cat_start = r
    for name, weight, keys in rb.CATEGORIES:
        sc.cell(row=r, column=1, value=name)
        sc.cell(row=r, column=2, value=weight).number_format = "0.0%"
        avg_formula = "=AVERAGE(" + ",".join(nm(k) for k in keys) + ")"
        sc.cell(row=r, column=3, value=avg_formula).number_format = "0.00"
        sc.cell(row=r, column=4, value=", ".join(k.replace("score_", "") for k in keys)).alignment = WRAP
        r += 1
    cat_end = r - 1

    # Weighted composite + market bonus
    r += 1
    sc.cell(row=r, column=1, value="Weighted avg (1-5)").font = Font(bold=True)
    sc.cell(row=r, column=3,
            value=f"=SUMPRODUCT(B{cat_start}:B{cat_end},C{cat_start}:C{cat_end})"
            ).number_format = "0.00"
    weighted_row = r
    r += 1
    sc.cell(row=r, column=1, value="Market-size bonus (TAM)").font = Font(italic=True)
    sc.cell(row=r, column=3,
            value=f"=MAX(-1,MIN(1,LOG10(MAX(1,{nm('tam_usd')})/1000000000)))*0.15"
            ).number_format = "0.00"
    bonus_row = r
    r += 1
    sc.cell(row=r, column=1, value="Composite (1-5)").font = Font(bold=True)
    sc.cell(row=r, column=3,
            value=f"=MAX(1,MIN(5,C{weighted_row}+C{bonus_row}))").number_format = "0.00"
    comp_row = r
    r += 1
    sc.cell(row=r, column=1, value="Score (0-100)").font = Font(bold=True)
    sc.cell(row=r, column=3, value=f"=(C{comp_row}-1)/4*100").number_format = "0.0"
    score100_row = r
    r += 1
    sc.cell(row=r, column=1, value="Verdict").font = Font(bold=True)
    sc.cell(row=r, column=3,
            value=(f'=IF(C{score100_row}>=75,"STRONG GO",'
                   f'IF(C{score100_row}>=60,"GO",'
                   f'IF(C{score100_row}>=45,"CONDITIONAL",'
                   f'IF(C{score100_row}>=30,"WEAK","NO-GO"))))'))

    # ---------------- Org sheet ----------------
    org = wb.create_sheet("Org")
    org["A1"] = "AI-Proof Org Structure"
    org["A1"].font = Font(bold=True, size=14, color="1F4E78")
    _header_row(org, 3,
                ["Function", "Role", "Level", "HC Y1", "HC Y3",
                 "Automation %", "Reports to", "AI augmentation",
                 "Cost Y1 (USD)", "Cost Y3 (USD)"],
                widths=[14, 26, 7, 9, 9, 13, 18, 60, 14, 14])

    roles = org_mod.generate(inputs)
    # We re-derive HC from named ranges so HC live-updates with Inputs sheet.
    y1_w = sum(t[3] for t in org_mod.TEMPLATE) or 1.0
    y3_w = sum(t[4] for t in org_mod.TEMPLATE) or 1.0

    r = 4
    for (fn, role, lvl, w1, w3, autom, reports, ai), gen in zip(org_mod.TEMPLATE, roles):
        org.cell(row=r, column=1, value=fn)
        org.cell(row=r, column=2, value=role)
        org.cell(row=r, column=3, value=lvl)
        # HC from named ranges (target_headcount_y1, _y3)
        org.cell(row=r, column=4,
                 value=f"=ROUND({w1}/{y1_w}*{nm('target_headcount_y1')},1)")
        org.cell(row=r, column=5,
                 value=f"=ROUND({w3}/{y3_w}*{nm('target_headcount_y3')},1)")
        # Effective automation blends template with user's target
        org.cell(row=r, column=6,
                 value=f"=ROUND({autom}*0.7+{nm('ai_automation_target')}*0.3,2)"
                 ).number_format = "0%"
        org.cell(row=r, column=7, value=reports)
        org.cell(row=r, column=8, value=ai).alignment = WRAP
        org.cell(row=r, column=9,
                 value=f"=D{r}*{nm('avg_fully_loaded_cost')}").number_format = "#,##0"
        org.cell(row=r, column=10,
                 value=f"=E{r}*{nm('avg_fully_loaded_cost')}").number_format = "#,##0"
        r += 1
    org.cell(row=r, column=3, value="TOTAL").font = Font(bold=True)
    org.cell(row=r, column=4, value=f"=SUM(D4:D{r-1})").number_format = "0.0"
    org.cell(row=r, column=5, value=f"=SUM(E4:E{r-1})").number_format = "0.0"
    org.cell(row=r, column=9, value=f"=SUM(I4:I{r-1})").number_format = "#,##0"
    org.cell(row=r, column=10, value=f"=SUM(J4:J{r-1})").number_format = "#,##0"

    # ---------------- DCF sheet ----------------
    dcf = wb.create_sheet("Valuation_DCF")
    dcf["A1"] = "DCF Valuation"
    dcf["A1"].font = Font(bold=True, size=14, color="1F4E78")
    _header_row(dcf, 3, ["Item"] + [f"Y{i}" for i in range(1, 6)],
                widths=[28, 14, 14, 14, 14, 14])
    dcf.cell(row=4, column=1, value="Revenue")
    dcf.cell(row=4, column=2, value=f"={nm('rev_y1')}").number_format = "#,##0"
    for i, g in enumerate(["rev_growth_y2", "rev_growth_y3",
                           "rev_growth_y4", "rev_growth_y5"], start=3):
        dcf.cell(row=4, column=i, value=f"={get_column_letter(i-1)}4*(1+{nm(g)})").number_format = "#,##0"
    dcf.cell(row=5, column=1, value="Gross profit")
    for i in range(2, 7):
        col = get_column_letter(i)
        dcf.cell(row=5, column=i, value=f"={col}4*{nm('gross_margin')}").number_format = "#,##0"
    dcf.cell(row=6, column=1, value="Opex")
    for i in range(2, 7):
        col = get_column_letter(i)
        dcf.cell(row=6, column=i, value=f"={col}4*{nm('opex_pct_rev')}").number_format = "#,##0"
    dcf.cell(row=7, column=1, value="EBIT")
    for i in range(2, 7):
        col = get_column_letter(i)
        dcf.cell(row=7, column=i, value=f"={col}5-{col}6").number_format = "#,##0"
    dcf.cell(row=8, column=1, value="NOPAT")
    for i in range(2, 7):
        col = get_column_letter(i)
        dcf.cell(row=8, column=i, value=f"={col}7*(1-{nm('tax_rate')})").number_format = "#,##0"
    dcf.cell(row=9, column=1, value="CapEx")
    for i in range(2, 7):
        col = get_column_letter(i)
        dcf.cell(row=9, column=i, value=f"={col}4*{nm('capex_pct_rev')}").number_format = "#,##0"
    dcf.cell(row=10, column=1, value="NWC")
    for i in range(2, 7):
        col = get_column_letter(i)
        dcf.cell(row=10, column=i, value=f"={col}4*{nm('nwc_pct_rev')}").number_format = "#,##0"
    dcf.cell(row=11, column=1, value="ΔNWC")
    dcf.cell(row=11, column=2, value="=B10").number_format = "#,##0"
    for i in range(3, 7):
        col = get_column_letter(i)
        prev = get_column_letter(i - 1)
        dcf.cell(row=11, column=i, value=f"={col}10-{prev}10").number_format = "#,##0"
    dcf.cell(row=12, column=1, value="FCF")
    for i in range(2, 7):
        col = get_column_letter(i)
        dcf.cell(row=12, column=i, value=f"={col}8-{col}9-{col}11").number_format = "#,##0"
    dcf.cell(row=13, column=1, value="Discount factor")
    for i in range(2, 7):
        period = i - 1
        dcf.cell(row=13, column=i, value=f"=1/(1+{nm('discount_rate')})^{period}").number_format = "0.000"
    dcf.cell(row=14, column=1, value="PV of FCF")
    for i in range(2, 7):
        col = get_column_letter(i)
        dcf.cell(row=14, column=i, value=f"={col}12*{col}13").number_format = "#,##0"

    dcf.cell(row=16, column=1, value="Terminal value").font = Font(bold=True)
    dcf.cell(row=16, column=2,
             value=f"=F12*(1+{nm('terminal_growth')})/({nm('discount_rate')}-{nm('terminal_growth')})"
             ).number_format = "#,##0"
    dcf.cell(row=17, column=1, value="PV of TV")
    dcf.cell(row=17, column=2,
             value=f"=B16/(1+{nm('discount_rate')})^5").number_format = "#,##0"
    dcf.cell(row=18, column=1, value="Enterprise Value").font = Font(bold=True)
    dcf.cell(row=18, column=2, value="=SUM(B14:F14)+B17").number_format = "#,##0"
    dcf.cell(row=19, column=1, value="Equity Value").font = Font(bold=True)
    dcf.cell(row=19, column=2, value=f"=B18-{nm('net_debt_usd')}").number_format = "#,##0"

    # ---------------- Comparables ----------------
    cmp = wb.create_sheet("Valuation_Comps")
    cmp["A1"] = "Comparables Valuation"
    cmp["A1"].font = Font(bold=True, size=14, color="1F4E78")
    cmp.column_dimensions["A"].width = 32
    cmp.column_dimensions["B"].width = 18
    cmp.cell(row=3, column=1, value="By revenue multiple")
    cmp.cell(row=3, column=2,
             value=f"={nm('comp_rev_multiple')}*{nm('comp_ntm_revenue')}").number_format = "#,##0"
    cmp.cell(row=4, column=1, value="By EBITDA multiple")
    cmp.cell(row=4, column=2,
             value=f"={nm('comp_ebitda_multiple')}*{nm('comp_ntm_ebitda')}").number_format = "#,##0"
    cmp.cell(row=5, column=1, value="Midpoint").font = Font(bold=True)
    cmp.cell(row=5, column=2, value="=AVERAGE(B3:B4)").number_format = "#,##0"

    # ---------------- Early-stage ----------------
    es = wb.create_sheet("Valuation_EarlyStage")
    es["A1"] = "Early-Stage Valuations"
    es["A1"].font = Font(bold=True, size=14, color="1F4E78")
    es.column_dimensions["A"].width = 36
    es.column_dimensions["B"].width = 18
    es.cell(row=3, column=1, value="Berkus total").font = Font(bold=True)
    es.cell(row=3, column=2,
            value=(f"={nm('berkus_idea')}+{nm('berkus_prototype')}+{nm('berkus_team')}"
                   f"+{nm('berkus_relationships')}+{nm('berkus_sales')}")).number_format = "#,##0"
    es.cell(row=4, column=1, value="Scorecard (Payne) pre-money").font = Font(bold=True)
    # multiplier = composite_5 / 3 (cross-sheet ref to Scorecard composite)
    es.cell(row=4, column=2,
            value=f"={nm('scorecard_avg_premoney')}*Scorecard!C{comp_row}/3").number_format = "#,##0"
    es.cell(row=5, column=1, value="VC method pre-money").font = Font(bold=True)
    es.cell(row=5, column=2,
            value=f"={nm('vc_exit_value')}/{nm('vc_target_roi')}").number_format = "#,##0"
    es.cell(row=6, column=1, value="First Chicago — success scenario")
    es.cell(row=6, column=2,
            value=f"=MAX(Valuation_DCF!B19,Valuation_Comps!B5)*1.5").number_format = "#,##0"
    es.cell(row=7, column=1, value="First Chicago — base scenario")
    es.cell(row=7, column=2,
            value=f"=AVERAGE(Valuation_DCF!B19,Valuation_Comps!B5)").number_format = "#,##0"
    es.cell(row=8, column=1, value="First Chicago — downside scenario")
    es.cell(row=8, column=2, value="=B3").number_format = "#,##0"
    es.cell(row=9, column=1, value="First Chicago — weighted").font = Font(bold=True)
    es.cell(row=9, column=2, value="=0.25*B6+0.5*B7+0.25*B8").number_format = "#,##0"

    # ---------------- Patent ----------------
    pt = wb.create_sheet("Valuation_Patent")
    pt["A1"] = "Patent / IP Valuation"
    pt["A1"].font = Font(bold=True, size=14, color="1F4E78")
    pt.column_dimensions["A"].width = 36
    pt.column_dimensions["B"].width = 18
    pt.cell(row=3, column=1, value="Cost approach (R&D + filing)").font = Font(bold=True)
    pt.cell(row=3, column=2, value=f"={nm('patent_dev_cost_usd')}").number_format = "#,##0"
    pt.cell(row=4, column=1, value="Market approach (comparable sale)").font = Font(bold=True)
    pt.cell(row=4, column=2, value=f"={nm('patent_comparable_sale_usd')}").number_format = "#,##0"
    pt.cell(row=5, column=1, value="Relief-from-royalty (PV)").font = Font(bold=True)
    pt.cell(row=5, column=2,
            value=(f"={nm('patent_royalty_rate')}*{nm('patent_revenue_attributable')}"
                   f"*(1-(1+{nm('discount_rate')})^-{nm('patent_useful_life_years')})"
                   f"/{nm('discount_rate')}")).number_format = "#,##0"
    pt.cell(row=6, column=1, value="Income approach (20% margin, PV)").font = Font(bold=True)
    pt.cell(row=6, column=2,
            value=(f"=0.20*{nm('patent_revenue_attributable')}"
                   f"*(1-(1+{nm('discount_rate')})^-{nm('patent_useful_life_years')})"
                   f"/{nm('discount_rate')}")).number_format = "#,##0"
    pt.cell(row=7, column=1, value="Midpoint of approaches").font = Font(bold=True)
    pt.cell(row=7, column=2, value="=AVERAGE(B3:B6)").number_format = "#,##0"

    # ---------------- Summary ----------------
    su = wb.create_sheet("Summary", 0)  # first sheet
    su.column_dimensions["A"].width = 36
    su.column_dimensions["B"].width = 24
    su["A1"] = f"{inputs.get('idea_name','Idea')} — Summary"
    su["A1"].font = Font(bold=True, size=16, color="1F4E78")
    su["A2"] = inputs.get("one_liner", "")
    su["A2"].font = Font(italic=True)
    su["A4"] = "Composite score (1-5)"; su["B4"] = f"=Scorecard!C{comp_row}"
    su["B4"].number_format = "0.00"
    su["A5"] = "Score (0-100)"; su["B5"] = f"=Scorecard!C{score100_row}"
    su["B5"].number_format = "0.0"
    su["A6"] = "Verdict"; su["B6"] = f"=Scorecard!C{score100_row+1}"
    su["A8"] = "DCF Equity Value (USD)"; su["B8"] = "=Valuation_DCF!B19"
    su["B8"].number_format = "#,##0"
    su["A9"] = "Comparables Midpoint (USD)"; su["B9"] = "=Valuation_Comps!B5"
    su["B9"].number_format = "#,##0"
    su["A10"] = "Berkus (USD)"; su["B10"] = "=Valuation_EarlyStage!B3"
    su["B10"].number_format = "#,##0"
    su["A11"] = "Scorecard (USD)"; su["B11"] = "=Valuation_EarlyStage!B4"
    su["B11"].number_format = "#,##0"
    su["A12"] = "VC method (USD)"; su["B12"] = "=Valuation_EarlyStage!B5"
    su["B12"].number_format = "#,##0"
    su["A13"] = "First Chicago weighted"; su["B13"] = "=Valuation_EarlyStage!B9"
    su["B13"].number_format = "#,##0"
    su["A14"] = "Patent midpoint (USD)"; su["B14"] = "=Valuation_Patent!B7"
    su["B14"].number_format = "#,##0"
    su["A16"] = "Org cost Y1 (USD)"; su["B16"] = f"=Org!I{r}"
    su["B16"].number_format = "#,##0"
    su["A17"] = "Org cost Y3 (USD)"; su["B17"] = f"=Org!J{r}"
    su["B17"].number_format = "#,##0"

    qual = results.get("qualitative")
    if qual:
        su["A20"] = "Qualitative analysis (AI)"
        su["A20"].font = Font(bold=True)
        su["A21"] = qual
        su["A21"].alignment = WRAP
        su.row_dimensions[21].height = 200
        su.merge_cells("A21:F21")

    # ---------------- Equity split sheet ----------------
    eqs = wb.create_sheet("Equity")
    eqs.column_dimensions["A"].width = 22
    for col in "BCDEFGH":
        eqs.column_dimensions[col].width = 14
    eqs["A1"] = "Founder Equity Split (Slicing-Pie style)"
    eqs["A1"].font = Font(bold=True, size=14, color="1F4E78")
    eqs["A2"] = ("Units = time%·salary_forgone + capital + ip_score·100k + "
                 "role_criticality·50k. Edit Inputs sheet — this recomputes.")
    eqs["A2"].font = Font(italic=True, color="666666")

    _header_row(eqs, 4,
                ["Founder", "Time %", "Salary forgone",
                 "Capital", "IP score", "Role criticality",
                 "Units", "Final equity %"],
                widths=[22, 10, 14, 14, 10, 14, 14, 14])

    units_rows = []
    for i in range(1, eq.N_FOUNDERS + 1):
        r = 4 + i
        keys = eq.founder_field_keys(i)
        eqs.cell(row=r, column=1, value=f"={nm(keys['name'])}") \
            if False else eqs.cell(row=r, column=1, value=inputs.get(keys['name']))
        eqs.cell(row=r, column=2, value=f"={nm(keys['time_pct'])}").number_format = "0%"
        eqs.cell(row=r, column=3, value=f"={nm(keys['salary_forgone'])}").number_format = "#,##0"
        eqs.cell(row=r, column=4, value=f"={nm(keys['capital_usd'])}").number_format = "#,##0"
        eqs.cell(row=r, column=5, value=f"={nm(keys['ip_score'])}")
        eqs.cell(row=r, column=6, value=f"={nm(keys['role_criticality'])}")
        # Units
        units_formula = (f"=B{r}*C{r}+D{r}+E{r}*{eq.IP_UNIT_VALUE}"
                         f"+F{r}*{eq.ROLE_UNIT_VALUE}")
        eqs.cell(row=r, column=7, value=units_formula).number_format = "#,##0"
        units_rows.append(r)

    first_r = units_rows[0]
    last_r = units_rows[-1]
    total_row = last_r + 1
    eqs.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    eqs.cell(row=total_row, column=7,
             value=f"=SUM(G{first_r}:G{last_r})").number_format = "#,##0"

    # Final equity % per founder = (units / total) * (1 - ESOP)
    for r in units_rows:
        eqs.cell(row=r, column=8,
                 value=f"=IF($G${total_row}=0,0,G{r}/$G${total_row}*(1-{nm('esop_pool_pct')}))"
                 ).number_format = "0.00%"

    esop_row = total_row + 1
    eqs.cell(row=esop_row, column=1, value="ESOP pool").font = Font(bold=True)
    eqs.cell(row=esop_row, column=8, value=f"={nm('esop_pool_pct')}").number_format = "0.00%"
    check_row = esop_row + 1
    eqs.cell(row=check_row, column=1, value="Check (=100%)").font = Font(italic=True)
    eqs.cell(row=check_row, column=8,
             value=f"=SUM(H{first_r}:H{last_r})+H{esop_row}").number_format = "0.00%"

    # ---------------- Screening sheet (single-idea view) ----------------
    # Lets the user enter the 8 RICE-style inputs for *this* idea and see
    # Modified RICE, WSJF, CD3, EMV side-by-side. Pre-filled with sensible
    # defaults derived from the deep-eval inputs where possible.
    sc = wb.create_sheet("Screening")
    sc.column_dimensions["A"].width = 36
    sc.column_dimensions["B"].width = 18
    sc.column_dimensions["C"].width = 60
    sc["A1"] = "Portfolio Screening (this idea)"
    sc["A1"].font = Font(bold=True, size=14, color="1F4E78")
    sc["A2"] = "Edit yellow cells. Compare the four methods — divergence flags scoring sensitivity."
    sc["A2"].font = Font(italic=True, color="666666")

    # 8 inputs; defaults derived from deep-eval inputs
    screen_defaults = {
        "monetary_impact": float(inputs.get("som_usd", 1_000_000)),
        "adoption": 0.5,
        "business_critical": min(1.0, float(inputs.get("score_problem_severity", 3)) / 5),
        "ownership": 1,
        "risk_pos": min(1.0, float(inputs.get("score_team_execution", 3)) / 5),
        "time_of_replacement_yrs": 3,
        "time_to_commercialize_yrs": 2,
        "cost_to_develop": float(inputs.get("avg_fully_loaded_cost", 40_000))
                          * max(1, int(inputs.get("target_headcount_y1", 8))),
    }
    rows = [
        ("Monetary impact ($/yr if successful)", "monetary_impact",  "#,##0"),
        ("Adoption (0-1)",                       "adoption",         "0.00"),
        ("Business critical (0-1)",              "business_critical","0.00"),
        ("Ownership (0 ext, 1 internal)",        "ownership",        "0"),
        ("Risk PoS (0-1)",                       "risk_pos",         "0.00"),
        ("Time of replacement (yrs)",            "time_of_replacement_yrs",   "0.0"),
        ("Time to commercialize (yrs)",          "time_to_commercialize_yrs", "0.0"),
        ("Cost to develop ($)",                  "cost_to_develop",  "#,##0"),
    ]
    screen_addr = {}
    r0 = 4
    for i, (label, key, fmt) in enumerate(rows):
        sc.cell(row=r0 + i, column=1, value=label)
        c = sc.cell(row=r0 + i, column=2, value=screen_defaults[key])
        c.fill = PatternFill("solid", fgColor="FFF2CC")
        c.border = BORDER
        c.number_format = fmt
        screen_addr[key] = f"B{r0+i}"

    A = screen_addr
    # Modified RICE
    sc["A14"] = "Modified RICE"; sc["A14"].font = Font(bold=True)
    sc["B14"] = (f"=({A['monetary_impact']}*{A['adoption']}*{A['business_critical']}"
                 f"*{A['ownership']}*{A['risk_pos']}*{A['time_of_replacement_yrs']})"
                 f"/({A['time_to_commercialize_yrs']}*{A['cost_to_develop']})")
    sc["B14"].number_format = "0.0000"
    sc["C14"] = "Reach × Impact × Confidence / Effort (Samyama variant)"
    sc["C14"].font = Font(italic=True, color="666666")

    # WSJF
    sc["A15"] = "WSJF"; sc["A15"].font = Font(bold=True)
    sc["B15"] = (f"=(MIN(100,{A['monetary_impact']}*{A['adoption']}*{A['risk_pos']}/100000)"
                 f"+MAX(0,MIN(1,1/MAX({A['time_of_replacement_yrs']},0.5)))*100"
                 f"+{A['business_critical']}*100)/{A['time_to_commercialize_yrs']}")
    sc["B15"].number_format = "0.0"
    sc["C15"] = "(BV + TimeCriticality + RiskReduction) / Duration  [SAFe]"
    sc["C15"].font = Font(italic=True, color="666666")

    # CD3
    sc["A16"] = "CD3 ($/yr per yr)"; sc["A16"].font = Font(bold=True)
    sc["B16"] = (f"=({A['monetary_impact']}*{A['adoption']}*{A['risk_pos']})"
                 f"/{A['time_to_commercialize_yrs']}")
    sc["B16"].number_format = "#,##0"
    sc["C16"] = "Cost of Delay / Duration  [Reinertsen]"
    sc["C16"].font = Font(italic=True, color="666666")

    # EMV
    sc["A17"] = "EMV (USD)"; sc["A17"].font = Font(bold=True)
    sc["B17"] = (f"={A['risk_pos']}*({A['monetary_impact']}*{A['adoption']})"
                 f"*{A['time_of_replacement_yrs']}-{A['cost_to_develop']}")
    sc["B17"].number_format = "#,##0"
    sc["C17"] = "PoS × (annual value × life) − cost"
    sc["C17"].font = Font(italic=True, color="666666")

    # ---------------- Definitions sheet ----------------
    df = wb.create_sheet("Definitions")
    df.column_dimensions["A"].width = 26
    df.column_dimensions["B"].width = 60
    df.column_dimensions["C"].width = 60
    df["A1"] = "Definitions & Method Notes"
    df["A1"].font = Font(bold=True, size=14, color="1F4E78")
    _header_row(df, 3, ["Term", "Definition", "How it's used in this tool"])
    rr = 4
    for term, definition, usage in defs.GLOSSARY:
        df.cell(row=rr, column=1, value=term).font = Font(bold=True)
        df.cell(row=rr, column=2, value=definition).alignment = WRAP
        df.cell(row=rr, column=3, value=usage).alignment = WRAP
        rr += 1
    rr += 2
    df.cell(row=rr, column=1, value="Rubric ↔ Screening cross-map").font = SECTION
    rr += 1
    for row in defs.RUBRIC_TO_SCREENING:
        is_header = row == defs.RUBRIC_TO_SCREENING[0]
        for i, val in enumerate(row, 1):
            c = df.cell(row=rr, column=i, value=val)
            if is_header:
                c.font = HEADER; c.fill = HEADER_FILL
            else:
                c.alignment = WRAP
        rr += 1

    wb.save(path)


# =========================================================================
#  Standalone screening workbook (--screen mode, many ideas)
# =========================================================================
def write_screening_workbook(scored_rows: list, path: str) -> None:
    """scored_rows comes from screening.rank_table(...)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Screening"
    ws["A1"] = "Portfolio Screening"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E78")
    ws["A2"] = ("Four methods ranked side-by-side. rank_spread = max(rank) − min(rank); "
                "high spread = scoring-sensitive idea, dig deeper.")
    ws["A2"].font = Font(italic=True, color="666666")

    headers = ["Idea", "Mod. RICE", "WSJF", "CD3 ($/yr)", "EMV ($)",
               "R(RICE)", "R(WSJF)", "R(CD3)", "R(EMV)", "Rank spread", "Avg rank"]
    widths = [28, 14, 12, 16, 16, 9, 9, 9, 9, 12, 12]
    _header_row(ws, 4, headers, widths=widths)
    for i, r in enumerate(scored_rows, start=5):
        ws.cell(row=i, column=1, value=r["name"])
        ws.cell(row=i, column=2, value=r["modified_rice"]).number_format = "0.0000"
        ws.cell(row=i, column=3, value=r["wsjf"]).number_format = "0.0"
        ws.cell(row=i, column=4, value=r["cd3"]).number_format = "#,##0"
        ws.cell(row=i, column=5, value=r["emv"]).number_format = "#,##0"
        ws.cell(row=i, column=6, value=r["rank_rice"])
        ws.cell(row=i, column=7, value=r["rank_wsjf"])
        ws.cell(row=i, column=8, value=r["rank_cd3"])
        ws.cell(row=i, column=9, value=r["rank_emv"])
        ws.cell(row=i, column=10, value=r["rank_spread"])
        ws.cell(row=i, column=11, value=r["avg_rank"]).number_format = "0.00"

    # Definitions sheet here too
    df = wb.create_sheet("Definitions")
    df.column_dimensions["A"].width = 26
    df.column_dimensions["B"].width = 60
    df.column_dimensions["C"].width = 60
    df["A1"] = "Method Definitions"
    df["A1"].font = Font(bold=True, size=14, color="1F4E78")
    _header_row(df, 3, ["Term", "Definition", "When to trust this method"])
    rr = 4
    for term, definition, usage in defs.GLOSSARY[:8]:  # screening terms only
        df.cell(row=rr, column=1, value=term).font = Font(bold=True)
        df.cell(row=rr, column=2, value=definition).alignment = WRAP
        df.cell(row=rr, column=3, value=usage).alignment = WRAP
        rr += 1

    wb.save(path)
